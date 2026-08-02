#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Rehace, en Python, el análisis que el estudiante hace en la planilla.
Sirve para (a) conciliar las cifras citadas en los .md con el dataset real
y (b) quedar como control cruzado del docente."""
import csv, os, datetime as dt
from collections import Counter, defaultdict

P = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "archivos")

def num(s):
    s = s.strip()
    if not s: return None
    return float(s.replace(".", "").replace(",", "."))

ZONA_CLI, TIPO_CLI = {}, {}
CAT_PROD, DESC_PROD = {}, {}
from openpyxl import load_workbook
wb = load_workbook(os.path.join(P, "Cartera_clientes_2026.xlsx"))
for r in list(wb["CLIENTES"].values)[1:]:
    ZONA_CLI[r[0]] = r[3]; TIPO_CLI[r[0]] = r[2]
wb = load_workbook(os.path.join(P, "Catalogo_productos_2026.xlsx"))
for r in list(wb["PRODUCTOS"].values)[1:]:
    CAT_PROD[r[0]] = r[2]; DESC_PROD[r[0]] = r[1]
wb = load_workbook(os.path.join(P, "Nomina_vendedores_2026.xlsx"))
NOM_VEN = {r[0]: r[1] for r in list(wb["VENDEDORES"].values)[1:]}
ZONA_VEN = {r[0]: r[2] for r in list(wb["VENDEDORES"].values)[1:]}

ESCALA = [(80,15),(40,10),(20,6),(10,3),(0,0)]
def autorizado(c):
    for d,p in ESCALA:
        if c >= d: return p
    return 0

rows = list(csv.DictReader(open(os.path.join(P,"SIGMA_export_ventas_202606.csv"),
                                encoding="utf-8"), delimiter=";"))
print(f"filas de datos: {len(rows)}")

ids = Counter(r["id_operacion"] for r in rows)
dups = {k:v for k,v in ids.items() if v>1}
print(f"ids duplicados: {len(dups)}  filas de más: {sum(v-1 for v in dups.values())}  -> {sorted(dups)}")

vac_desc = sum(1 for r in rows if not r["descuento_pct"].strip())
neg = [r for r in rows if int(r["cantidad"])<0]
cero = [r for r in rows if num(r["precio_unitario"]) in (0.0,None)]
iso  = sum(1 for r in rows if "-" in r["fecha"])
est  = Counter(r["estado"] for r in rows)
prod_sucio = sum(1 for r in rows if r["cod_producto"] != r["cod_producto"].strip().upper())
print(f"descuento vacío: {vac_desc} | cantidad negativa: {len(neg)} | precio 0: {len(cero)}")
print(f"fechas ISO: {iso} | cod_producto sucio: {prod_sucio}")
print("estados:", dict(est))

huerf_cli = Counter(r["cod_cliente"] for r in rows if r["cod_cliente"] not in ZONA_CLI)
huerf_prod = Counter(r["cod_producto"].strip().upper() for r in rows
                     if r["cod_producto"].strip().upper() not in CAT_PROD)
print("clientes huérfanos:", dict(huerf_cli), "| productos huérfanos:", dict(huerf_prod))
sin_venta = [p for p in CAT_PROD if p not in
             {r["cod_producto"].strip().upper() for r in rows}]
print("productos sin ventas:", sin_venta)

def fecha(r):
    t = r["fecha"]
    return dt.date.fromisoformat(t) if "-" in t else dt.datetime.strptime(t,"%d/%m/%Y").date()
v7 = [r for r in rows if r["cod_vendedor"]=="VEN-07"]
v7_prev = [r["id_operacion"] for r in v7 if fecha(r) < dt.date(2026,6,15)]
print(f"VEN-07: {len(v7)} ops, {len(v7_prev)} anteriores al ingreso -> {v7_prev}")

# ---------------------------------------------------------------- importes
def ok_estado(r):
    return r["estado"].strip().upper().rstrip(".") in ("ENTREGADO","ENTREG")
def bruto(r):
    p = num(r["precio_unitario"]) or 0
    return int(r["cantidad"]) * p
def neto(r):
    d = num(r["descuento_pct"]) or 0
    return bruto(r) * (1 - d/100)

total_naive = sum(neto(r) for r in rows)                       # todo, con duplicados
vistos, limpio = set(), []
for r in rows:
    k = (r["id_operacion"],)
    if k in vistos: continue
    vistos.add(k); limpio.append(r)
util = [r for r in limpio if ok_estado(r) and int(r["cantidad"])>0
        and (num(r["precio_unitario"]) or 0) > 0]
total_limpio = sum(neto(r) for r in util)
print(f"\nTOTAL 'como lo informa el sistema' (todo, con duplicados): {total_naive:,.0f}")
print(f"TOTAL depurado (sin duplicados, solo entregadas válidas):   {total_limpio:,.0f}")
print(f"DIFERENCIA: {total_naive-total_limpio:,.0f}")
print(f"  de la cual duplicados: {sum(neto(r) for r in rows)-sum(neto(r) for r in limpio):,.0f}")
print(f"  anuladas/pendientes:   {sum(neto(r) for r in limpio if not ok_estado(r)):,.0f}")
print(f"  devoluciones (neg.):   {sum(neto(r) for r in limpio if ok_estado(r) and int(r['cantidad'])<0):,.0f}")
print(f"ticket promedio: {total_limpio/len(util):,.0f} sobre {len(util)} operaciones")

# ---------------------------------------------------------------- vendedor
OBJ = {"VEN-01":3_600_000,"VEN-02":2_900_000,"VEN-03":4_800_000,
       "VEN-04":2_400_000,"VEN-05":4_500_000,"VEN-06":2_000_000}
print("\nvendedor            facturación   objetivo   cumpl.")
fact = defaultdict(float)
for r in util: fact[r["cod_vendedor"]] += neto(r)
for v in sorted(fact):
    o = OBJ.get(v)
    c = f"{fact[v]/o*100:6.1f}%" if o else "   s/o"
    print(f"{v} {NOM_VEN.get(v,''):<18} {fact[v]:>11,.0f} {o or 0:>10,} {c}")

# ---------------------------------------------------------------- descuentos
exc = []
for r in util:
    d = num(r["descuento_pct"]) or 0
    a = autorizado(int(r["cantidad"]))
    if d > a:
        exc.append((r, d-a, bruto(r)*(d-a)/100))
print(f"\nexcesos de descuento: {len(exc)} operaciones, costo {sum(e[2] for e in exc):,.0f}")
print("  por vendedor:", dict(Counter(e[0]["cod_vendedor"] for e in exc)))

# ---------------------------------------------------------------- cortes
print("\npor zona:")
z = defaultdict(float); zc = defaultdict(set)
for r in util:
    zz = ZONA_CLI.get(r["cod_cliente"], "(sin zona)")
    z[zz] += neto(r); zc[zz].add(r["cod_cliente"])
for k in sorted(z, key=lambda x:-z[x]):
    print(f"  {k:<18} {z[k]:>11,.0f}  {z[k]/total_limpio*100:5.1f}%  clientes={len(zc[k])}")
print("por categoría:")
c = defaultdict(float)
for r in util: c[CAT_PROD.get(r["cod_producto"].strip().upper(),"(huérfano)")] += neto(r)
for k in sorted(c, key=lambda x:-c[x]):
    print(f"  {k:<22} {c[k]:>11,.0f}  {c[k]/total_limpio*100:5.1f}%")
print("por día de la semana:")
dsem = defaultdict(float); nom=["lunes","martes","miércoles","jueves","viernes"]
for r in util: dsem[fecha(r).weekday()] += neto(r)
for k in sorted(dsem): print(f"  {nom[k]:<12} {dsem[k]:>11,.0f}")
print("ticket promedio por zona:")
for k in sorted(z, key=lambda x:-z[x]):
    n = sum(1 for r in util if ZONA_CLI.get(r["cod_cliente"],"(sin zona)")==k)
    print(f"  {k:<18} {z[k]/n:>10,.0f}")
