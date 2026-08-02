#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Genera los archivos de práctica del Proyecto Andes Sur (Clases 08.1-11.1).

  archivos/SIGMA_export_ventas_202606.csv   <- export "sucio" del ERP
  archivos/Catalogo_productos_2026.xlsx     <- maestro de productos
  archivos/Cartera_clientes_2026.xlsx       <- maestro de clientes
  archivos/Nomina_vendedores_2026.xlsx      <- maestro de vendedores

Los defectos del CSV son deliberados y están documentados en la clase 08.1.
Semilla fija: el dataset es reproducible byte a byte.
"""
import csv, os, random, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

random.seed(20260731)

OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "archivos")
os.makedirs(OUT, exist_ok=True)

# ----------------------------------------------------------------- maestros

PRODUCTOS = [
    ("PRD-091", "Gaseosa cola 2,25 L",       "Bebidas sin alcohol", "Andina",        "Caja x6",  4980, 18),
    ("PRD-118", "Agua saborizada 1,5 L",     "Bebidas sin alcohol", "Aquavida",      "Caja x6",  8450, 22),
    ("PRD-204", "Cerveza rubia lata 473 ml", "Bebidas con alcohol", "Patagonia Sur", "Caja x12", 12300, 14),
    ("PRD-330", "Papas fritas 250 g",        "Snacks",              "Crocantina",    "Caja x10", 6700, 31),
    ("PRD-412", "Galletitas surtidas 400 g", "Almacén",             "Delicia",       "Caja x12", 9900, 26),
    ("PRD-505", "Yerba mate 1 kg",           "Almacén",             "Del Litoral",   "Caja x6",  15200, 19),
]
PRECIO = {p[0]: p[5] for p in PRODUCTOS}
# PRD-505 nunca se vende: es el hallazgo "producto sin rotación" de la clase 10.1.
VENDIBLES = ["PRD-091", "PRD-118", "PRD-204", "PRD-330", "PRD-412"]
PESO_PRODUCTO = [0.26, 0.24, 0.18, 0.20, 0.12]

CLIENTES = [
    ("CLI-1002", "Almacén Don Pedro",      "Almacén",              "Quilmes",          "Bernal",         "Monotributo",     400000),
    ("CLI-1007", "Autoservicio La Esquina","Minimercado",          "Quilmes",          "Quilmes Centro", "Resp. Inscripto", 1200000),
    ("CLI-1008", "Kiosco 24hs Ferrari",    "Kiosco",               "Avellaneda",       "Sarandí",        "Monotributo",     250000),
    ("CLI-1013", "Supermercado Alfa",      "Minimercado",          "Berazategui",      "Berazategui",    "Resp. Inscripto", 1800000),
    ("CLI-1015", "Distribuidora El Cruce", "Mayorista",            "Avellaneda",       "Villa Domínico", "Resp. Inscripto", 3000000),
    ("CLI-1019", "Almacén Sol",            "Almacén",              "Quilmes",          "Ezpeleta",       "Monotributo",     300000),
    ("CLI-1021", "Estación Norte SRL",     "Estación de servicio", "Florencio Varela", "Varela Centro",  "Resp. Inscripto", 900000),
    ("CLI-1027", "Minimercado Las Flores", "Minimercado",          "Avellaneda",       "Piñeyro",        "Monotributo",     500000),
    ("CLI-1032", "Kiosco Rivadavia",       "Kiosco",               "Quilmes",          "Bernal",         "Monotributo",     200000),
    ("CLI-1036", "Autoservicio Vega",      "Minimercado",          "Florencio Varela", "Bosques",        "Resp. Inscripto", 750000),
    ("CLI-1050", "Tienda Online Sur",      "E-commerce",           "Lanús",            "Lanús Oeste",    "Resp. Inscripto", 1500000),
]
ZONA_CLI = {c[0]: c[3] for c in CLIENTES}

VENDEDORES = [
    ("VEN-01", "Aguirre, Marina",  "Quilmes",          "03/02/2021", "Relación de dependencia"),
    ("VEN-02", "Benítez, Hugo",    "Berazategui",      "17/09/2019", "Relación de dependencia"),
    ("VEN-03", "Cardozo, Silvina", "Quilmes",          "01/03/2018", "Relación de dependencia"),
    ("VEN-04", "Duarte, Ramiro",   "Florencio Varela", "22/11/2022", "Monotributo"),
    ("VEN-05", "Espósito, Laura",  "Avellaneda",       "08/06/2020", "Relación de dependencia"),
    ("VEN-06", "Figueroa, Julián", "Lanús",            "14/04/2023", "Monotributo"),
    ("VEN-07", "Ferrari, Nicolás", "(sin asignar)",    "15/06/2026", "Monotributo"),
]
ZONA_VEN = {v[0]: v[2] for v in VENDEDORES}
CLI_POR_ZONA = {}
for c in CLIENTES:
    CLI_POR_ZONA.setdefault(c[3], []).append(c[0])

CANALES = ["Preventa", "Telefónico", "WhatsApp", "E-commerce"]
PESO_CANAL = [0.52, 0.20, 0.22, 0.06]

ESCALA = [(80, 15), (40, 10), (20, 6), (10, 3), (0, 0)]
def autorizado(cant):
    for desde, pct in ESCALA:
        if cant >= desde:
            return pct
    return 0

# ----------------------------------------------------------------- fechas

def dias_habiles():
    d, out = dt.date(2026, 6, 1), []
    while d.month == 6:
        if d.weekday() < 5:
            out.append(d)
        d += dt.timedelta(days=1)
    return out
DIAS = dias_habiles()
# La distribución no es pareja a propósito: lunes y jueves son los días
# fuertes de reparto. Es el hallazgo de "agrupar por día de la semana".
PESO_DIA = {0: 1.55, 1: 0.85, 2: 0.80, 3: 1.45, 4: 1.00}

# ----------------------------------------------------------------- filas base
# Las 20 primeras operaciones son exactamente las publicadas en la Guía
# Visual de la clase 08.1: el estudiante tiene que poder cargarlas a mano
# y encontrar las mismas filas en el archivo completo.

SEMILLA = [
    ("OP-10021","01/06/2026","VEN-03","CLI-1007","PRD-118",12,"8.450,00","3","Preventa","Entregado"),
    ("OP-10022","01/06/2026","VEN-01","CLI-1032","PRD-204",6,"12.300,00","","Telefónico","Entregado"),
    ("OP-10023","2026-06-02","VEN-05","CLI-1015","PRD-118",24,"8.450,00","6","Preventa","ENTREGADO"),
    ("OP-10023","2026-06-02","VEN-05","CLI-1015","PRD-118",24,"8.450,00","6","Preventa","ENTREGADO"),
    ("OP-10024","02/06/2026","VEN-02","CLI-1044","PRD-091",10,"4.980,00","0","WhatsApp","entregado"),
    ("OP-10025","02/06/2026","VEN-03","CLI-1007","prd-118",8,"8.450,00","0","Preventa","Pendiente"),
    ("OP-10026","03/06/2026","VEN-04","CLI-1021","PRD-330",15,"6.700,00","3","Preventa","Entregado"),
    ("OP-10027","03/06/2026","VEN-01","CLI-1032","PRD-204",-6,"12.300,00","","Telefónico","Entregado"),
    ("OP-10028","03/06/2026","VEN-06","CLI-1050","PRD-091",20,"4.980,00","6","E-commerce","Anulado"),
    ("OP-10029","04/06/2026","VEN-02","CLI-1013","PRD-118 ",9,"8.450,00","0","WhatsApp","Entreg."),
    ("OP-10030","04/06/2026","VEN-05","CLI-1015","PRD-412",30,"0","0","Preventa","Entregado"),
    ("OP-10031","2026-06-05","VEN-07","CLI-1008","PRD-330",14,"6.700,00","3","Preventa","Entregado"),
    ("OP-10032","05/06/2026","VEN-04","CLI-1021","PRD-204",4,"12.300,00","0","Telefónico","pendiente"),
    ("OP-10033","05/06/2026","VEN-03","CLI-1019","PRD-091",18,"4.980,00","","Preventa","Entregado"),
    ("OP-10034","08/06/2026","VEN-06","CLI-1050","PRD-412",25,"9.900,00","15","E-commerce","Entregado"),
    ("OP-10035","08/06/2026","VEN-01","CLI-1002","PRD-118",11,"8.450,00","3","Preventa","Entregado"),
    ("OP-10036","09/06/2026","VEN-02","CLI-1013","PRD-330",7,"6.700,00","0","WhatsApp","Entregado"),
    ("OP-10037","09/06/2026","VEN-05","CLI-1027","PRD-204",13,"12.300,00","3","Preventa","Entregado"),
    ("OP-10038","10/06/2026","VEN-03","CLI-1007","PRD-412",16,"9.900,00","3","Preventa","Entregado"),
    ("OP-10039","10/06/2026","VEN-04","CLI-1036","PRD-091",5,"4.980,00","","Telefónico","Anulado"),
]

# ----------------------------------------------------------------- helpers

def ar(n):
    """4980 -> '4.980,00' (formato de importe del export, texto)."""
    e = f"{int(n):,}".replace(",", ".")
    return f"{e},00"

def fecha_txt(d, iso=False):
    return d.isoformat() if iso else d.strftime("%d/%m/%Y")

def elegir(seq, pesos):
    return random.choices(seq, weights=pesos, k=1)[0]

def dia_al_azar(desde=None):
    cand = [d for d in DIAS if desde is None or d >= desde]
    return random.choices(cand, weights=[PESO_DIA[d.weekday()] for d in cand], k=1)[0]

# Cuánto vende cada vendedor (peso relativo). VEN-03 y VEN-05 son los
# pesados; VEN-06 (Lanús, un solo cliente) el más chico; VEN-07 entró
# el 15/06 y solo trabaja media quincena.
PESO_VEN = {"VEN-01": 1.05, "VEN-02": 0.92, "VEN-03": 1.42,
            "VEN-04": 0.86, "VEN-05": 1.30, "VEN-06": 0.55, "VEN-07": 0.30}

def cantidad(cliente=None):
    """Cajas por operación: sesgada a valores bajos, con cola larga.
    El cliente de Lanús compra siempre chico — es el dato que sostiene el
    análisis de zona de la clase 11.1."""
    if cliente == "CLI-1050":
        return random.randint(2, 9)
    r = random.random()
    if r < 0.52:  return random.randint(1, 5)
    if r < 0.82:  return random.randint(6, 11)
    if r < 0.95:  return random.randint(12, 22)
    if r < 0.99:  return random.randint(23, 45)
    return random.randint(46, 85)

# ----------------------------------------------------------------- generación

filas = list(SEMILLA)
nid = 10040
TOTAL = 406   # + 6 duplicados insertados abajo = 412 filas de datos

# operaciones de VEN-07 anteriores a su fecha de ingreso (hallazgo de
# control interno de la clase 10.1): 3 en total, una ya en la semilla.
ven07_previas = 2

while len(filas) < TOTAL:
    ven = elegir(list(PESO_VEN), list(PESO_VEN.values()))

    # fecha
    if ven == "VEN-07":
        if ven07_previas > 0 and random.random() < 0.35:
            d = random.choice([x for x in DIAS if x < dt.date(2026, 6, 15)])
            ven07_previas -= 1
        else:
            d = dia_al_azar(dt.date(2026, 6, 15))
    else:
        d = dia_al_azar()

    # cliente: casi siempre de la zona del vendedor
    zv = ZONA_VEN[ven]
    if ven == "VEN-07" or zv not in CLI_POR_ZONA or random.random() < 0.09:
        cli = random.choice([c[0] for c in CLIENTES])          # zona cruzada
    else:
        cli = random.choice(CLI_POR_ZONA[zv])
    if random.random() < 0.012:
        cli = "CLI-1044"                                        # huérfano

    prod = elegir(VENDIBLES, PESO_PRODUCTO)
    if random.random() < 0.008:
        prod = "PRD-660"                                        # huérfano
    cant = cantidad(cli)

    # canal: el e-commerce es casi todo del cliente de Lanús
    canal = "E-commerce" if cli == "CLI-1050" and random.random() < 0.75 \
            else elegir(CANALES[:3], PESO_CANAL[:3])

    # descuento: dentro de la escala. Los excesos se aplican después,
    # sobre la cantidad definitiva (ver más abajo).
    aut = autorizado(cant)
    if random.random() < 0.13:
        desc = ""                                               # sin cargar
    else:
        desc = max(0, aut - random.choice([0, 0, 0, 1, 2]))

    # estado
    r = random.random()
    if   r < 0.845: estado = random.choice(["Entregado"] * 12 + ["ENTREGADO", "entregado", "Entreg."])
    elif r < 0.935: estado = random.choice(["Pendiente"] * 5 + ["pendiente"])
    else:           estado = "Anulado"

    precio = PRECIO.get(prod, 7300)
    if random.random() < 0.007:
        precio = 0                                              # registro incompleto
    if random.random() < 0.014:
        cant = -abs(random.randint(2, 9))                        # devolución sin marcar

    # suciedad de formato
    pcod = prod
    r = random.random()
    if   r < 0.012: pcod = prod.lower()
    elif r < 0.022: pcod = prod + " "
    fecha = fecha_txt(d, iso=(random.random() < 0.07))

    filas.append((f"OP-{nid}", fecha, ven, cli, pcod, cant,
                  ar(precio) if precio else "0",
                  str(desc) if desc != "" else "", canal, estado))
    nid += 1

# ordenar por fecha real (el export sale ordenado por fecha de carga)
def clave(f):
    t = f[1]
    d = dt.date.fromisoformat(t) if "-" in t else dt.datetime.strptime(t, "%d/%m/%Y").date()
    return (d, f[0])
filas.sort(key=clave)

# --- excesos de descuento -------------------------------------------------
# Exactamente 11 operaciones por encima de la escala autorizada, 8 de ellas
# del mismo vendedor: es el hallazgo que Dirección eleva en la semana 4, y
# tiene que ser nítido para que la recomendación del informe se sostenga.
# OP-10034 (VEN-06, 25 cajas al 15%) ya viene con exceso en la semilla
# publicada en la Guía, así que acá se cargan las 10 restantes.
filas = [list(f) for f in filas]
SEM_IDS = {f[0] for f in SEMILLA}

def es_valida(f):
    return (f[9].strip().upper().rstrip(".") in ("ENTREGADO", "ENTREG")
            and int(f[5]) > 0 and f[6] != "0")

CUPO_EXCESO = {"VEN-04": 8, "VEN-01": 1, "VEN-05": 1}
for ven, n in CUPO_EXCESO.items():
    elegibles = [i for i, f in enumerate(filas)
                 if f[2] == ven and f[0] not in SEM_IDS and es_valida(f)
                 and int(f[5]) >= 8]
    for i in random.sample(elegibles, min(n, len(elegibles))):
        cant = int(filas[i][5])
        filas[i][7] = str(autorizado(cant) + random.choice([4, 5, 6, 7, 8]))

# duplicados por doble carga: 6 más, además de OP-10023 que ya viene
# duplicada en la semilla. Se insertan pegados al original.
ya = {f[0] for f in filas if f[0] == "OP-10023"}
candidatos = [i for i, f in enumerate(filas) if f[0] != "OP-10023"]
for i in sorted(random.sample(candidatos, 6), reverse=True):
    filas.insert(i + 1, filas[i])

# ----------------------------------------------------------------- CSV

path_csv = os.path.join(OUT, "SIGMA_export_ventas_202606.csv")
with open(path_csv, "w", encoding="utf-8", newline="") as fh:
    w = csv.writer(fh, delimiter=";", lineterminator="\r\n")
    w.writerow(["id_operacion", "fecha", "cod_vendedor", "cod_cliente",
                "cod_producto", "cantidad", "precio_unitario",
                "descuento_pct", "canal", "estado"])
    for f in filas:
        w.writerow(f)

# ----------------------------------------------------------------- XLSX

HDR_FILL = PatternFill("solid", fgColor="1E8E3E")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

def xlsx(nombre, hoja, cols, datos, anchos):
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(cols)
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HDR_FILL, HDR_FONT
        cell.alignment = Alignment(vertical="center")
    for r in datos:
        ws.append(list(r))
    for i, a in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = a
    ws.freeze_panes = "A2"
    p = os.path.join(OUT, nombre)
    wb.save(p)
    return p

xlsx("Catalogo_productos_2026.xlsx", "PRODUCTOS",
     ["cod_producto", "descripcion", "categoria", "marca", "unidad_venta",
      "precio_lista", "margen_pct"],
     PRODUCTOS, [14, 30, 22, 16, 14, 14, 13])

xlsx("Cartera_clientes_2026.xlsx", "CLIENTES",
     ["cod_cliente", "razon_social", "tipo_comercio", "zona", "localidad",
      "condicion_iva", "limite_credito"],
     CLIENTES, [13, 26, 22, 18, 18, 18, 16])

xlsx("Nomina_vendedores_2026.xlsx", "VENDEDORES",
     ["cod_vendedor", "apellido_nombre", "zona_asignada", "fecha_ingreso",
      "modalidad"],
     VENDEDORES, [15, 22, 18, 15, 26])

print(f"CSV: {len(filas)} filas + encabezado -> {path_csv}")
